"""
Indexer — Scans repositories and indexes them into Qdrant.
Supports incremental mode: only re-indexes changed files.
Run with: docker compose run --rm indexer
"""

import os
import sys
import glob
import time
import hashlib

import openpyxl
import xlrd

from qdrant_client import QdrantClient
from qdrant_client.models import (
    Distance, VectorParams, PointStruct,
    PayloadSchemaType, TextIndexParams, TokenizerType,
    HnswConfigDiff, ScalarQuantization, ScalarQuantizationConfig, ScalarType,
    Filter, FieldCondition, MatchValue,
)
from sentence_transformers import SentenceTransformer

QDRANT_URL = os.getenv("QDRANT_URL", "http://localhost:6333")
COLLECTION_NAME = os.getenv("COLLECTION_NAME", "codebase")
REPOS_PATH = os.getenv("REPOS_PATH", "./repos")
CHUNK_MAX_CHARS = int(os.getenv("CHUNK_MAX_CHARS", "1500"))
CHUNK_OVERLAP_CHARS = int(os.getenv("CHUNK_OVERLAP_CHARS", "200"))
EMBEDDING_MODEL = os.getenv("EMBEDDING_MODEL", "BAAI/bge-base-en-v1.5")
BATCH_SIZE = int(os.getenv("BATCH_SIZE", "64"))
HNSW_M = int(os.getenv("HNSW_M", "16"))
HNSW_EF = int(os.getenv("HNSW_EF", "128"))
ENABLE_QUANTIZATION = os.getenv("ENABLE_QUANTIZATION", "true").lower() == "true"
FORCE_REINDEX = os.getenv("FORCE_REINDEX", "false").lower() == "true"

FILE_EXTENSIONS = [
    "*.py", "*.js", "*.ts", "*.tsx", "*.jsx",
    "*.java", "*.kt", "*.scala", "*.cs", "*.vb",
    "*.go", "*.rs", "*.rb", "*.php",
    "*.c", "*.cpp", "*.h", "*.hpp", "*.swift", "*.m",
    "*.sql",
    "*.md", "*.txt", "*.rst",
    "*.yaml", "*.yml", "*.toml", "*.json",
    "*.xlsx", "*.xls",
]

SKIP_DIRS = {
    "node_modules", ".git", "__pycache__", "venv", ".venv",
    "env", ".env", "dist", "build", ".next", ".nuxt",
    "target", "bin", "obj", ".idea", ".vscode",
    "vendor", "packages", ".tox", "coverage",
    "test_data", "fixtures", "migrations",
}

SKIP_FILES = {
    "package-lock.json", "yarn.lock", "pnpm-lock.yaml",
    "Cargo.lock", "poetry.lock", "composer.lock",
    ".DS_Store", "Thumbs.db",
}

MAX_FILE_SIZE = 100_000


def _get_overlap_lines(lines: list[str], max_overlap: int) -> tuple[list[str], int]:
    """Returns (overlap_lines, overlap_length) from the end of the given lines."""
    if max_overlap <= 0 or not lines:
        return [], 0
    overlap_lines = []
    overlap_len = 0
    for line in reversed(lines):
        overlap_len += len(line) + 1
        overlap_lines.append(line)
        if overlap_len >= max_overlap:
            break
    overlap_lines.reverse()
    return overlap_lines, overlap_len


def chunk_codice(testo: str, filepath: str, max_chars: int = CHUNK_MAX_CHARS,
                 overlap: int = CHUNK_OVERLAP_CHARS) -> list[str]:
    header = f"# File: {filepath}\n"
    righe = testo.split("\n")
    chunks, chunk_corrente, lunghezza = [], [], 0

    block_starters = (
        "def ", "class ", "function ", "async function ",
        "public ", "private ", "protected ", "internal ",
        "export ", "const ", "let ", "var ",
        "## ", "# ", "### ",
        "func ", "fn ", "impl ",
        "interface ", "struct ", "enum ",
        "module ", "namespace ",
        "describe(", "it(", "test(",
        "router.get(", "router.post(", "router.put(", "router.delete(", "router.patch(",
        "app.get(", "app.post(", "app.put(", "app.delete(", "app.use(",
        "module.exports",
    )

    for riga in righe:
        if any(riga.strip().startswith(s) for s in block_starters) and lunghezza > max_chars // 3:
            chunks.append(header + "\n".join(chunk_corrente))
            overlap_lines, overlap_len = _get_overlap_lines(chunk_corrente, overlap)
            chunk_corrente, lunghezza = overlap_lines, overlap_len

        chunk_corrente.append(riga)
        lunghezza += len(riga) + 1

        if lunghezza >= max_chars:
            chunks.append(header + "\n".join(chunk_corrente))
            overlap_lines, overlap_len = _get_overlap_lines(chunk_corrente, overlap)
            chunk_corrente, lunghezza = overlap_lines, overlap_len

    if chunk_corrente:
        chunks.append(header + "\n".join(chunk_corrente))

    return [c for c in chunks if len(c.strip()) > 50]


def read_xls(filepath: str) -> str:
    """Extract text from a legacy .xls file using xlrd."""
    wb = xlrd.open_workbook(filepath)
    parts = []
    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        rows = []
        for row_idx in range(ws.nrows):
            row = ws.row_values(row_idx)
            if all(v == "" or v is None for v in row):
                continue
            rows.append("\t".join("" if v is None else str(v) for v in row))
        if rows:
            parts.append(f"## Sheet: {sheet_name}\n" + "\n".join(rows))
    return "\n\n".join(parts)


def read_xlsx(filepath: str) -> str:
    """Extract text from an xlsx file. Each sheet is rendered as a markdown-like table."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                continue
            rows.append("\t".join("" if v is None else str(v) for v in row))
        if rows:
            parts.append(f"## Sheet: {sheet_name}\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts)


def chunk_xlsx(testo: str, filepath: str, max_chars: int = CHUNK_MAX_CHARS,
               overlap: int = CHUNK_OVERLAP_CHARS) -> list[str]:  # noqa: ARG001
    """Chunk xlsx text: split by sheet boundaries first, then by rows."""
    header = f"# File: {filepath}\n"
    chunks = []
    sheets = testo.split("\n\n## Sheet: ")
    for i, section in enumerate(sheets):
        section_text = section if i == 0 else "## Sheet: " + section
        lines = section_text.split("\n")
        current, length = [], 0
        for line in lines:
            current.append(line)
            length += len(line) + 1
            if length >= max_chars:
                chunks.append(header + "\n".join(current))
                sheet_header = [ln for ln in current if ln.startswith("## Sheet:")]
                current = sheet_header
                length = sum(len(ln) + 1 for ln in current)
        if current:
            chunks.append(header + "\n".join(current))
    return [c for c in chunks if len(c.strip()) > 50]


def should_skip(filepath: str) -> bool:
    parts = filepath.split(os.sep)
    if any(d in SKIP_DIRS for d in parts):
        return True
    if os.path.basename(filepath) in SKIP_FILES:
        return True
    try:
        return os.path.getsize(filepath) > MAX_FILE_SIZE
    except OSError:
        return True


def _content_hash(content: str) -> str:
    """MD5 hash of file content for change detection."""
    return hashlib.md5(content.encode()).hexdigest()


def _chunk_id(rel_path: str, chunk_index: int) -> str:
    return hashlib.md5(f"{rel_path}::chunk_{chunk_index}".encode()).hexdigest()


def scan_repos(repos_path: str) -> list[dict]:
    if not os.path.exists(repos_path):
        print(f"[ERROR] Folder not found: {repos_path}")
        sys.exit(1)

    repo_dirs = sorted([d for d in os.listdir(repos_path) if os.path.isdir(os.path.join(repos_path, d)) and not d.startswith(".")])
    if not repo_dirs:
        print(f"[ERROR] No repositories in: {repos_path}")
        print(f"   Check REPOS_HOST_PATH in .env")
        sys.exit(1)

    documenti = []
    for repo_name in repo_dirs:
        repo_path = os.path.join(repos_path, repo_name)
        file_count, chunk_count = 0, 0

        for ext in FILE_EXTENSIONS:
            for filepath in glob.glob(os.path.join(repo_path, "**", ext), recursive=True):
                if should_skip(filepath):
                    continue
                file_ext = os.path.splitext(filepath)[1].lower()
                try:
                    if file_ext == ".xlsx":
                        contenuto = read_xlsx(filepath)
                        chunker = chunk_xlsx
                    elif file_ext == ".xls":
                        contenuto = read_xls(filepath)
                        chunker = chunk_xlsx
                    else:
                        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                            contenuto = f.read()
                        chunker = chunk_codice
                except Exception:
                    continue
                if not contenuto.strip() or len(contenuto) < 20:
                    continue

                file_count += 1
                rel_path = os.path.relpath(filepath, repos_path)
                file_hash = _content_hash(contenuto)
                for i, chunk in enumerate(chunker(contenuto, rel_path)):
                    documenti.append({
                        "id": _chunk_id(rel_path, i),
                        "content": chunk,
                        "file": rel_path,
                        "repo": repo_name,
                        "chunk_index": i,
                        "extension": file_ext,
                        "content_hash": file_hash,
                    })
                    chunk_count += 1

        print(f"  {repo_name}: {file_count} files -> {chunk_count} chunks")
    return documenti


def _collection_exists(client: QdrantClient) -> bool:
    collections = [c.name for c in client.get_collections().collections]
    return COLLECTION_NAME in collections


def _create_collection(client: QdrantClient, vector_size: int):
    print(f"  Creating collection: {COLLECTION_NAME} (dim={vector_size})")
    quantization = ScalarQuantization(
        scalar=ScalarQuantizationConfig(type=ScalarType.INT8, always_ram=True)
    ) if ENABLE_QUANTIZATION else None
    client.create_collection(
        collection_name=COLLECTION_NAME,
        vectors_config=VectorParams(size=vector_size, distance=Distance.COSINE),
        hnsw_config=HnswConfigDiff(m=HNSW_M, ef_construct=HNSW_EF),
        quantization_config=quantization,
    )

    # Indexes for filters and full-text search
    client.create_payload_index(collection_name=COLLECTION_NAME, field_name="repo", field_schema=PayloadSchemaType.KEYWORD)
    client.create_payload_index(collection_name=COLLECTION_NAME, field_name="extension", field_schema=PayloadSchemaType.KEYWORD)
    client.create_payload_index(collection_name=COLLECTION_NAME, field_name="file", field_schema=PayloadSchemaType.KEYWORD)
    client.create_payload_index(collection_name=COLLECTION_NAME, field_name="content_hash", field_schema=PayloadSchemaType.KEYWORD)
    client.create_payload_index(
        collection_name=COLLECTION_NAME, field_name="content",
        field_schema=TextIndexParams(type="text", tokenizer=TokenizerType.WORD, min_token_len=2, max_token_len=20, lowercase=True),
    )


def _get_existing_hashes(client: QdrantClient) -> dict[str, set[str]]:
    """Get a map of file -> content_hash for all indexed points."""
    file_hashes = {}
    offset = None
    while True:
        results, offset = client.scroll(
            collection_name=COLLECTION_NAME,
            limit=1000,
            offset=offset,
            with_payload=["file", "content_hash", "chunk_index"],
            with_vectors=False,
        )
        for point in results:
            file_path = point.payload.get("file", "")
            content_hash = point.payload.get("content_hash", "")
            if file_path and content_hash:
                file_hashes[file_path] = content_hash
        if offset is None:
            break
    return file_hashes


def _delete_file_points(client: QdrantClient, file_path: str):
    """Delete all points belonging to a specific file."""
    client.delete(
        collection_name=COLLECTION_NAME,
        points_selector=Filter(must=[FieldCondition(key="file", match=MatchValue(value=file_path))]),
    )


def indicizza(documenti: list[dict], force: bool = False):
    print(f"\n  Qdrant: {QDRANT_URL}")
    client = QdrantClient(url=QDRANT_URL)

    print(f"  Embedding: {EMBEDDING_MODEL}")
    embedder = SentenceTransformer(EMBEDDING_MODEL, model_kwargs={"attn_implementation": "eager"})
    vector_size = embedder.get_sentence_embedding_dimension()

    exists = _collection_exists(client)

    if force or not exists:
        # Full reindex
        if exists:
            print(f"  Drop collection: {COLLECTION_NAME}")
            client.delete_collection(COLLECTION_NAME)
        _create_collection(client, vector_size)
        docs_to_index = documenti
        print(f"  Full index: {len(docs_to_index)} chunks")
    else:
        # Incremental: only index changed files
        print("  Incremental mode: checking for changes...")
        existing_hashes = _get_existing_hashes(client)

        # Group new docs by file
        new_files = {}
        for doc in documenti:
            file_path = doc["file"]
            if file_path not in new_files:
                new_files[file_path] = {"hash": doc["content_hash"], "docs": []}
            new_files[file_path]["docs"].append(doc)

        # Find changed/new/deleted files
        changed_files = set()
        new_file_set = set(new_files.keys())
        existing_file_set = set(existing_hashes.keys())

        # Files that are new or changed
        for file_path, info in new_files.items():
            if file_path not in existing_hashes or existing_hashes[file_path] != info["hash"]:
                changed_files.add(file_path)

        # Files that were deleted from repos
        deleted_files = existing_file_set - new_file_set

        if not changed_files and not deleted_files:
            print("  No changes detected. Skipping.")
            return client.count(COLLECTION_NAME).count

        # Delete old points for changed + deleted files
        for file_path in changed_files | deleted_files:
            _delete_file_points(client, file_path)

        if deleted_files:
            print(f"  Removed {len(deleted_files)} deleted files")

        # Only index docs from changed files
        docs_to_index = []
        for file_path in changed_files:
            docs_to_index.extend(new_files[file_path]["docs"])

        print(f"  Changed: {len(changed_files)} files -> {len(docs_to_index)} chunks to index")

    # Batch encode and upsert
    total = len(docs_to_index)
    if total == 0:
        return client.count(COLLECTION_NAME).count

    for start in range(0, total, BATCH_SIZE):
        end = min(start + BATCH_SIZE, total)
        batch = docs_to_index[start:end]
        embeddings = embedder.encode([d["content"] for d in batch], show_progress_bar=False)
        points = [
            PointStruct(id=d["id"], vector=emb.tolist(), payload={
                "content": d["content"], "file": d["file"], "repo": d["repo"],
                "chunk_index": d["chunk_index"], "extension": d["extension"],
                "content_hash": d["content_hash"],
            })
            for d, emb in zip(batch, embeddings)
        ]
        client.upsert(collection_name=COLLECTION_NAME, points=points)
        print(f"    {min(100, int(end / total * 100))}% ({end}/{total})")

    return client.count(COLLECTION_NAME).count


def main():
    print("=" * 60)
    print("RAG Indexer")
    print("=" * 60)
    print(f"  Repos: {REPOS_PATH}")
    print(f"  Qdrant: {QDRANT_URL}")
    print(f"  Chunk: max={CHUNK_MAX_CHARS} chars, overlap={CHUNK_OVERLAP_CHARS} chars")
    print(f"  HNSW: m={HNSW_M} ef={HNSW_EF}, quantization={'on' if ENABLE_QUANTIZATION else 'off'}")
    print(f"  Mode: {'full (forced)' if FORCE_REINDEX else 'incremental'}")
    print()

    t = time.time()
    docs = scan_repos(REPOS_PATH)
    if not docs:
        print("[ERROR] No documents found!")
        sys.exit(1)

    print(f"\n  Total: {len(docs)} chunks scanned")
    n = indicizza(docs, force=FORCE_REINDEX)
    print(f"\n  Completed in {time.time() - t:.1f}s — {n} vectors in collection")
    print("=" * 60)


if __name__ == "__main__":
    main()
